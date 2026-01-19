"""
设计文件生成器 - Excel文档生成

职责：
1. 打开设计文件模板
2. 写入所有行（封面+目录+图纸）
3. 导出PDF

依赖：
- openpyxl: Excel操作
- 参数规范.yaml: design_bindings配置

测试要点：
- test_generate_design_file: 设计文件生成
- test_design_global_fields: 全局字段写入
- test_design_frame_fields: 图纸行字段
- test_design_cover_catalog_rows: 封面/目录行特化
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from ..config import load_spec
from ..interfaces import GenerationError, IDesignFileGenerator
from .pdf_engine import PDFExporter

if TYPE_CHECKING:
    from ..models import DocContext


class DesignFileGenerator(IDesignFileGenerator):
    """设计文件生成器实现"""
    
    def __init__(
        self, 
        spec_path: str | None = None,
        pdf_exporter: PDFExporter | None = None,
    ):
        self.spec = load_spec(spec_path) if spec_path else load_spec()
        self.pdf_exporter = pdf_exporter or PDFExporter()
    
    def generate(self, ctx: DocContext, output_dir: Path) -> tuple[Path, Path]:
        """生成设计文件"""
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # 1. 获取模板路径
        template_path = self.spec.get_template_path("design", ctx.params.project_no)
        if not Path(template_path).exists():
            raise GenerationError(f"设计文件模板不存在: {template_path}")
        
        # 2. 获取落点配置
        bindings = self.spec.get_design_bindings()
        
        # 3. 写入Excel
        output_xlsx = output_dir / "设计文件.xlsx"
        self._write_design(template_path, output_xlsx, bindings, ctx)
        
        # 4. 导出PDF
        output_pdf = output_dir / "设计文件.pdf"
        self.pdf_exporter.export_xlsx_to_pdf(output_xlsx, output_pdf)
        
        return output_xlsx, output_pdf
    
    def _write_design(
        self, 
        template_path: str, 
        output_path: Path,
        bindings: dict,
        ctx: DocContext,
    ) -> None:
        """写入设计文件Excel"""
        wb = load_workbook(template_path)
        ws = wb.active
        
        start_row = bindings.get("start_row", 2)
        columns = bindings.get("columns", {})
        
        # 准备全局数据
        global_data = self._prepare_global_data(ctx)
        
        # 行顺序：封面 → 目录 → 图纸
        rows = self._build_rows(ctx)
        
        current_row = start_row
        for row_data in rows:
            self._write_row(ws, current_row, row_data, global_data, columns, ctx)
            current_row += 1
        
        wb.save(output_path)
    
    def _prepare_global_data(self, ctx: DocContext) -> dict:
        """准备全局数据（所有行相同）"""
        params = ctx.params
        derived = ctx.derived
        mappings = self.spec.get_mappings()
        
        # 专业代码映射
        discipline_code = mappings.get("discipline_to_code", {}).get(
            params.discipline or "", ""
        )
        
        return {
            "design_status": params.design_status,
            "wbs_code": params.wbs_code,
            "album_internal_code": derived.album_internal_code,
            "internal_tag": params.internal_tag,
            "subitem_name": params.subitem_name,
            "subitem_no": params.subitem_no,
            "system_code": params.system_code,
            "system_name": params.system_name,
            "discipline": params.discipline,
            "discipline_code": discipline_code,
            "discipline_office": params.discipline_office,
            "design_phase": derived.design_phase,
            "classification": params.classification,
            "file_category": params.file_category,
            "attachment_name": params.attachment_name,
            "qa_required": params.qa_required,
            "qa_engineer": params.qa_engineer,
            "work_hours": params.work_hours,
        }
    
    def _build_rows(self, ctx: DocContext) -> list[dict]:
        """构建行数据"""
        rows = []
        derived = ctx.derived
        params = ctx.params
        
        # 封面行
        rows.append({
            "type": "cover",
            "external_code": derived.cover_external_code,
            "internal_code": derived.cover_internal_code,
            "revision": params.cover_revision,
            "title_cn": derived.cover_title_cn,
            "title_en": derived.cover_title_en,
            "paper_size_text": "A4文件",
            "page_total": 1,
            "status": params.doc_status,
        })
        
        # 目录行
        rows.append({
            "type": "catalog",
            "external_code": derived.catalog_external_code,
            "internal_code": derived.catalog_internal_code,
            "revision": derived.catalog_revision,
            "title_cn": derived.catalog_title_cn,
            "title_en": derived.catalog_title_en,
            "paper_size_text": "A4文件",
            "page_total": derived.catalog_page_total or 1,
            "status": params.doc_status,
        })
        
        # 图纸行
        for frame in ctx.get_sorted_frames():
            tb = frame.titleblock
            rows.append({
                "type": "drawing",
                "external_code": tb.external_code,
                "internal_code": tb.internal_code,
                "revision": tb.revision,
                "title_cn": tb.title_cn,
                "title_en": tb.title_en,
                "paper_size_text": tb.paper_size_text,
                "page_total": tb.page_total or 1,
                "status": tb.status,
            })
        
        return rows
    
    def _write_row(
        self, 
        ws, 
        row: int, 
        row_data: dict, 
        global_data: dict,
        columns: dict,
        ctx: DocContext,
    ) -> None:
        """写入单行"""
        # 遍历列配置写入
        for col_letter, col_config in columns.items():
            source = col_config.get("source", "")
            is_global = col_config.get("global", False)
            
            # 确定值
            if is_global:
                value = global_data.get(source, "")
            elif source in row_data:
                value = row_data[source]
            else:
                value = ""
            
            # 写入
            ws[f"{col_letter}{row}"] = value
