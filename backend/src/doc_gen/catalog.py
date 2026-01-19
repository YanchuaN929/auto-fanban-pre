"""
目录生成器 - Excel文档生成

职责：
1. 打开目录模板
2. 写入表头和明细行
3. 计算页数（优先Excel分页信息，兜底PDF计页）
4. 回填页数后导出PDF

依赖：
- openpyxl: Excel操作
- 参数规范.yaml: catalog_bindings配置

测试要点：
- test_generate_catalog_common: 通用目录生成
- test_generate_catalog_1818: 1818目录（中英文标题同格）
- test_catalog_row_order: 行顺序（封面→目录→图纸）
- test_catalog_page_count: 页数计算
- test_catalog_upgrade_note: 升版标记
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from ..config import load_spec
from ..interfaces import GenerationError, ICatalogGenerator
from .pdf_engine import PDFExporter

if TYPE_CHECKING:
    from ..models import DocContext, FrameMeta


class CatalogGenerator(ICatalogGenerator):
    """目录生成器实现"""
    
    def __init__(
        self, 
        spec_path: str | None = None,
        pdf_exporter: PDFExporter | None = None,
    ):
        self.spec = load_spec(spec_path) if spec_path else load_spec()
        self.pdf_exporter = pdf_exporter or PDFExporter()
    
    def generate(self, ctx: DocContext, output_dir: Path) -> tuple[Path, Path, int]:
        """生成目录文档"""
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # 1. 选择模板
        template_path = self._get_template_path(ctx)
        if not Path(template_path).exists():
            raise GenerationError(f"目录模板不存在: {template_path}")
        
        # 2. 获取落点配置
        bindings = self.spec.get_catalog_bindings()
        
        # 3. 写入Excel
        output_xlsx = output_dir / "目录.xlsx"
        self._write_catalog(template_path, output_xlsx, bindings, ctx)
        
        # 4. 计算页数（优先Excel分页信息）
        page_count = self._count_pages(output_xlsx)
        
        # 5. 回填目录行页数
        self._backfill_page_count(output_xlsx, page_count, bindings)
        
        # 6. 导出PDF
        output_pdf = output_dir / "目录.pdf"
        self.pdf_exporter.export_xlsx_to_pdf(output_xlsx, output_pdf)
        
        return output_xlsx, output_pdf, page_count
    
    def _get_template_path(self, ctx: DocContext) -> str:
        """获取模板路径"""
        return self.spec.get_template_path("catalog", ctx.params.project_no)
    
    def _write_catalog(
        self, 
        template_path: str, 
        output_path: Path,
        bindings: dict,
        ctx: DocContext,
    ) -> None:
        """写入目录Excel"""
        wb = load_workbook(template_path)
        ws = wb.active
        
        # 写入表头
        self._write_header(ws, bindings, ctx)
        
        # 写入明细行
        start_row = bindings.get("detail", {}).get("start_row", 9)
        current_row = start_row
        
        # 行顺序：封面 → 目录 → 图纸（按internal_code尾号升序）
        rows = self._build_detail_rows(ctx)
        
        for row_data in rows:
            self._write_detail_row(ws, current_row, row_data, bindings, ctx)
            current_row += 1
        
        # 保存
        wb.save(output_path)
    
    def _write_header(self, ws, bindings: dict, ctx: DocContext) -> None:
        """写入表头"""
        header = bindings.get("header", {})
        derived = ctx.derived
        params = ctx.params
        
        # engineering_no → C1
        if "engineering_no" in header:
            cell = header["engineering_no"].get("cell", "C1")
            ws[cell] = params.engineering_no
        
        # catalog_internal_code → H1
        if "catalog_internal_code" in header:
            cell = header["catalog_internal_code"].get("cell", "H1")
            ws[cell] = derived.catalog_internal_code
        
        # catalog_external_code → H3
        if "catalog_external_code" in header:
            cell = header["catalog_external_code"].get("cell", "H3")
            ws[cell] = derived.catalog_external_code
        
        # subitem_no → C5
        if "subitem_no" in header:
            cell = header["subitem_no"].get("cell", "C5")
            ws[cell] = params.subitem_no
        
        # catalog_revision → H5
        if "catalog_revision" in header:
            cell = header["catalog_revision"].get("cell", "H5")
            ws[cell] = derived.catalog_revision
    
    def _build_detail_rows(self, ctx: DocContext) -> list[dict]:
        """构建明细行数据"""
        rows = []
        derived = ctx.derived
        params = ctx.params
        
        # 1. 封面行
        rows.append({
            "type": "cover",
            "internal_code": derived.cover_internal_code,
            "external_code": derived.cover_external_code,
            "title_cn": derived.cover_title_cn,
            "title_en": derived.cover_title_en,
            "revision": params.cover_revision,
            "status": params.doc_status,
            "page_total": 1,
            "upgrade_note": "",
        })
        
        # 2. 目录行
        rows.append({
            "type": "catalog",
            "internal_code": derived.catalog_internal_code,
            "external_code": derived.catalog_external_code,
            "title_cn": derived.catalog_title_cn,
            "title_en": derived.catalog_title_en,
            "revision": derived.catalog_revision,
            "status": params.doc_status,
            "page_total": 0,  # 占位，后续回填
            "upgrade_note": "",
        })
        
        # 3. 图纸行（按internal_code尾号升序）
        for frame in ctx.get_sorted_frames():
            tb = frame.titleblock
            seq_no = tb.get_seq_no()
            
            # 判断是否需要升版标记
            upgrade_note = ""
            if (
                params.upgrade_start_seq is not None 
                and params.upgrade_end_seq is not None
                and seq_no is not None
            ):
                if params.upgrade_start_seq <= seq_no <= params.upgrade_end_seq:
                    upgrade_note = params.upgrade_note_text
            
            rows.append({
                "type": "drawing",
                "internal_code": tb.internal_code,
                "external_code": tb.external_code,
                "title_cn": tb.title_cn,
                "title_en": tb.title_en,
                "revision": tb.revision,
                "status": tb.status,
                "page_total": tb.page_total or 1,
                "upgrade_note": upgrade_note,
            })
        
        return rows
    
    def _write_detail_row(
        self, 
        ws, 
        row: int, 
        data: dict, 
        bindings: dict,
        ctx: DocContext,
    ) -> None:
        """写入单行明细"""
        columns = bindings.get("detail", {}).get("columns", {})
        
        # A: 序号
        ws[f"A{row}"] = row - bindings.get("detail", {}).get("start_row", 9) + 1
        
        # B: 图纸编号（internal_code）
        if "B" in columns:
            ws[f"B{row}"] = data.get("internal_code", "")
        
        # D: 文件编码（external_code）
        if "D" in columns:
            ws[f"D{row}"] = data.get("external_code", "")
        
        # E: 名称（1818需要中英文换行）
        if "E" in columns:
            title = data.get("title_cn", "")
            if ctx.is_1818 and data.get("title_en"):
                title = f"{title}\n{data['title_en']}"
                ws[f"E{row}"].alignment = ws[f"E{row}"].alignment.copy(wrapText=True)
            ws[f"E{row}"] = title
        
        # F: 版次
        if "F" in columns:
            ws[f"F{row}"] = data.get("revision", "")
        
        # G: 状态
        if "G" in columns:
            ws[f"G{row}"] = data.get("status", "")
        
        # H: 页数
        if "H" in columns:
            ws[f"H{row}"] = data.get("page_total", 1)
        
        # I: 附注（升版标记）
        if "I" in columns:
            ws[f"I{row}"] = data.get("upgrade_note", "")
    
    def _count_pages(self, xlsx_path: Path) -> int:
        """计算目录页数"""
        # 优先尝试Excel分页信息
        try:
            wb = load_workbook(xlsx_path)
            ws = wb.active
            
            # 尝试通过分页符计算
            h_breaks = len(ws.page_breaks.horizontalBreaks) if hasattr(ws, 'page_breaks') else 0
            if h_breaks > 0:
                return h_breaks + 1
        except Exception:
            pass
        
        # 兜底：导出PDF计页
        try:
            temp_pdf = xlsx_path.with_suffix(".temp.pdf")
            self.pdf_exporter.export_xlsx_to_pdf(xlsx_path, temp_pdf)
            count = self.pdf_exporter.count_pdf_pages(temp_pdf)
            temp_pdf.unlink(missing_ok=True)
            return count
        except Exception:
            return 1  # 默认1页
    
    def _backfill_page_count(
        self, 
        xlsx_path: Path, 
        page_count: int,
        bindings: dict,
    ) -> None:
        """回填目录行页数"""
        wb = load_workbook(xlsx_path)
        ws = wb.active
        
        # 目录行是第2行明细（封面后）
        start_row = bindings.get("detail", {}).get("start_row", 9)
        catalog_row = start_row + 1
        
        ws[f"H{catalog_row}"] = page_count
        
        wb.save(xlsx_path)
