"""
IED计划生成器 - Excel文档生成（仅Excel，不导出PDF）

职责：
1. 打开IED计划模板
2. 写入所有行（封面+目录+图纸）
3. 单独输出Excel（不入package.zip）

依赖：
- openpyxl: Excel操作
- 参数规范.yaml: ied_bindings配置

测试要点：
- test_generate_ied: IED计划生成
- test_ied_columns: 列映射正确性
- test_ied_fixed_values: 固定值列
- test_ied_no_pdf: 不导出PDF
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from ..config import load_spec
from ..interfaces import GenerationError, IIEDGenerator

if TYPE_CHECKING:
    from ..models import DocContext


class IEDGenerator(IIEDGenerator):
    """IED计划生成器实现"""
    
    def __init__(self, spec_path: str | None = None):
        self.spec = load_spec(spec_path) if spec_path else load_spec()
    
    def generate(self, ctx: DocContext, output_dir: Path) -> Path:
        """生成IED计划（仅Excel）"""
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # 1. 获取模板路径
        template_path = self.spec.get_template_path("ied", ctx.params.project_no)
        if not Path(template_path).exists():
            raise GenerationError(f"IED计划模板不存在: {template_path}")
        
        # 2. 获取落点配置
        bindings = self.spec.get_ied_bindings()
        
        # 3. 写入Excel
        output_xlsx = output_dir / "IED计划.xlsx"
        self._write_ied(template_path, output_xlsx, bindings, ctx)
        
        # 注意：IED不导出PDF
        return output_xlsx
    
    def _write_ied(
        self, 
        template_path: str, 
        output_path: Path,
        bindings: dict,
        ctx: DocContext,
    ) -> None:
        """写入IED计划Excel"""
        wb = load_workbook(template_path)
        
        # 使用指定的sheet
        sheet_name = bindings.get("sheet", "IED导入模板 (修改)")
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        start_row = bindings.get("start_row", 2)
        columns = bindings.get("columns", {})
        
        # 准备全局数据
        global_data = self._prepare_global_data(ctx)
        
        # 行顺序：封面 → 目录 → 图纸
        rows = self._build_rows(ctx)
        
        current_row = start_row
        for row_data in rows:
            self._write_row(ws, current_row, row_data, global_data, columns, ctx)
            current_row += 1
        
        wb.save(output_path)
    
    def _prepare_global_data(self, ctx: DocContext) -> dict:
        """准备全局数据"""
        params = ctx.params
        derived = ctx.derived
        
        return {
            "ied_change_flag": params.ied_change_flag,
            "ied_doc_type": params.ied_doc_type,
            "ied_status": params.ied_status,
            "wbs_code": params.wbs_code,
            "album_internal_code": derived.album_internal_code,
            "classification": params.classification,
            "work_hours": params.work_hours,
            # ... 其他IED参数
        }
    
    def _build_rows(self, ctx: DocContext) -> list[dict]:
        """构建行数据"""
        rows = []
        derived = ctx.derived
        params = ctx.params
        
        # 封面行
        rows.append({
            "type": "cover",
            "external_code": derived.cover_external_code,
            "internal_code": derived.cover_internal_code,
            "revision": params.cover_revision,
            "title_cn": derived.cover_title_cn,
            "title_en": derived.cover_title_en,
        })
        
        # 目录行
        rows.append({
            "type": "catalog",
            "external_code": derived.catalog_external_code,
            "internal_code": derived.catalog_internal_code,
            "revision": derived.catalog_revision,
            "title_cn": derived.catalog_title_cn,
            "title_en": derived.catalog_title_en,
        })
        
        # 图纸行
        for frame in ctx.get_sorted_frames():
            tb = frame.titleblock
            rows.append({
                "type": "drawing",
                "external_code": tb.external_code,
                "internal_code": tb.internal_code,
                "revision": tb.revision,
                "title_cn": tb.title_cn,
                "title_en": tb.title_en,
            })
        
        return rows
    
    def _write_row(
        self, 
        ws, 
        row: int, 
        row_data: dict, 
        global_data: dict,
        columns: dict,
        ctx: DocContext,
    ) -> None:
        """写入单行"""
        for col_letter, col_config in columns.items():
            # 固定值
            if "value" in col_config:
                ws[f"{col_letter}{row}"] = col_config["value"]
                continue
            
            source = col_config.get("source", "")
            is_global = col_config.get("global", False)
            
            if is_global:
                value = global_data.get(source, "")
            elif source in row_data:
                value = row_data[source]
            else:
                value = ""
            
            ws[f"{col_letter}{row}"] = value
